#!/usr/bin/env python3
"""
yaml2excel.py

Multi-environment Helm values.yaml comparator:
- Input: a JSON file mapping environment name -> path to values.yaml
    e.g. compare.json
    {
        "dev": "./dev/values.yaml",
        "sit": "./sit/values.yaml",
        "uat": "./uat/values.yaml",
        "prod": "./prod/values.yaml"
    }
- Usage:
    python yaml2excel.py compare.json
    python yaml2excel.py compare.json my_compare_name.xlsx

- Output: Excel file with columns: Key | <env1> | <env2> | ...
  Differences across envs are highlighted in yellow.

Requirements:
    pip install -r requirements.txt
    (PyYAML, pandas, openpyxl)
"""

from collections import OrderedDict
import json
import sys
import os
from typing import Any, Dict, List, Tuple
import yaml
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# ---------- Configuration ----------
YAML_SAFE_LOAD = yaml.safe_load
HIGHLIGHT_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # yellow
HEADER_FONT = Font(bold=True)
# -----------------------------------

def row_has_difference(row):
    values = list(row)
    # If all are NaN → no difference
    if all(pd.isna(v) for v in values):
        return False
    # Compare each value against the first
    first_val = values[0]
    for val in values[1:]:
        # Case 1: one is NaN, other is not
        if pd.isna(val) != pd.isna(first_val):
            return True
        # Case 2: values are not equal
        if not pd.isna(val) and val != first_val:
            return True
    return False


def usage_and_exit(msg: str = None, code: int = 1):
    if msg:
        print(f"Error: {msg}")
    print(
        "Usage:\n"
        "  python yaml2excel.py compare.json [output.xlsx]\n\n"
        "  - compare.json: JSON file mapping env -> path to values.yaml\n"
        "  - output.xlsx (optional): explicit Excel filename. If omitted,\n"
        "    output will be auto-generated as: compare_<env1>_<env2>....xlsx"
    )
    sys.exit(code)


def load_json_config(path: str) -> "OrderedDict[str, str]":
    if not os.path.isfile(path):
        raise FileNotFoundError(f"Config JSON file not found: {path}")
    try:
        with open(path, "r", encoding="utf-8") as f:
            # Preserve order explicitly (json.load preserves order in Python 3.7+, but be explicit)
            data = json.load(f, object_pairs_hook=OrderedDict)
            if not isinstance(data, dict) or len(data) < 2:
                raise ValueError("Config JSON must be an object with at least 2 environment mappings.")
            # Convert to OrderedDict of str->str and validate
            ordered = OrderedDict()
            for k, v in data.items():
                if not isinstance(k, str):
                    raise ValueError("Environment names in JSON must be strings.")
                if not isinstance(v, str):
                    raise ValueError("YAML file paths in JSON must be strings.")
                ordered[k] = v
            return ordered
    except json.JSONDecodeError as e:
        raise ValueError(f"Invalid JSON file {path}: {e}")


def read_yaml_file(path: str) -> Any:
    if not os.path.isfile(path):
        raise FileNotFoundError(f"YAML file not found: {path}")
    try:
        with open(path, "r", encoding="utf-8") as f:
            content = YAML_SAFE_LOAD(f)  # returns dict, list, or scalar, or None
            return content if content is not None else {}
    except yaml.YAMLError as e:
        raise ValueError(f"Invalid YAML in file {path}: {e}")


def flatten_yaml(
    data: Any,
    parent_key: str = "",
    sep: str = "."
) -> List[Tuple[str, Any]]:
    """
    Flatten YAML structure preserving traversal order.
    Returns list of (flattened_key, value) in the order they are encountered.
    - dict -> iterate items() (python preserves insertion order)
    - list -> indexes like [0]
    - scalar -> assign to parent_key
    """
    items: List[Tuple[str, Any]] = []

    if isinstance(data, dict):
        for k, v in data.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else k
            items.extend(flatten_yaml(v, new_key, sep=sep))
    elif isinstance(data, list):
        for i, v in enumerate(data):
            new_key = f"{parent_key}[{i}]"
            items.extend(flatten_yaml(v, new_key, sep=sep))
    else:
        # Scalars (including None)
        items.append((parent_key, data))
    return items


def build_master_key_order(list_of_env_flat_items: List[List[Tuple[str, Any]]]) -> List[str]:
    """
    Build master key order: keys from first env (in order), then any new keys from subsequent envs in their
    discovered order appended at the end (first-seen basis).
    """
    seen = set()
    order: List[str] = []
    for flat_items in list_of_env_flat_items:
        for key, _ in flat_items:
            if key not in seen:
                seen.add(key)
                order.append(key)
    return order


def flatten_items_to_dict(flat_items: List[Tuple[str, Any]]) -> Dict[str, Any]:
    """
    Convert list of (key, val) to dict. If duplicate keys occur within the same file (rare), last wins.
    """
    return {k: v for k, v in flat_items}


def build_dataframe(master_keys: List[str], env_names: List[str], env_key_value_maps: Dict[str, Dict[str, Any]]) -> pd.DataFrame:
    """
    Construct DataFrame with columns: Key, <env1>, <env2>, ...
    Fill missing values with empty string.
    """
    rows = []
    for key in master_keys:
        row = {"Key": key}
        for env in env_names:
            val = env_key_value_maps.get(env, {}).get(key, "")
            # Convert Python None -> empty string, keep scalars as-is (int, bool, str)
            if val is None:
                val = ""
            row[env] = val
        rows.append(row)
    df = pd.DataFrame(rows, columns=["Key"] + env_names)
    return df


def autogen_output_filename(env_names: List[str]) -> str:
    safe_envs = [e.replace(" ", "_") for e in env_names]
    filename = "compare_" + "_".join(safe_envs) + ".xlsx"
    return filename


def write_excel_with_highlight(df: pd.DataFrame, output_path: str, env_names: List[str]) -> None:
    """
    Write DataFrame to Excel and then highlight differing cells (yellow) using openpyxl.
    Highlight rule:
      - For each row, gather non-empty values across env columns.
      - If more than one unique non-empty value exists -> highlight all env cells in that row.
    """
    # Write with pandas first
    try:
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="compare", index=False)
    except Exception as e:
        raise IOError(f"Failed to write Excel file {output_path}: {e}")

    # Load workbook to apply styles
    wb = load_workbook(output_path)
    ws = wb["compare"]

    # Style header row
    header_row = 1
    for col_idx in range(1, len(df.columns) + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = HEADER_FONT

    # Map env columns to excel column indexes: Key is col 1, envs start at col 2
    env_col_indexes = {env: idx + 2 for idx, env in enumerate(env_names)}

    # Iterate rows starting from row 2
    for r in range(2, ws.max_row + 1):
        values = []
        # collect non-empty string representations for comparison
        for env in env_names:
            cell = ws.cell(row=r, column=env_col_indexes[env])
            # Use value as-is for comparison; convert Excel None -> ""
            val = cell.value
            if val is None:
                val = ""
            # Important: normalize booleans and numbers to string for uniform comparison?
            # We'll compare by their string representation, but keep original in cell.
            values.append(val)
        non_blank_vals = [v for v in values if v != ""]
        # Determine unique set (use string form for robust comparison)
        unique_vals = set()
        for v in values:
          if v == "" or v is None:
             unique_vals.add("__MISSING__") #Channged made to highlight missing values 
          else:   
            unique_vals.add(str(v).strip())
        if len(unique_vals) > 1:
             for env in env_names:
                col = env_col_indexes[env]
                cell = ws.cell(row=r, column=col)
                # highlight only non-empty cells, but per requirement we can highlight all env cells (including blanks)
                # We will highlight non-blank cells; if you prefer highlighting blanks too, remove the condition.
                # Here we'll highlight all env cells that are not equal across envs (so non-blank or blank if others differ).
                cell.fill = HIGHLIGHT_FILL

    # Save workbook
    try:
        wb.save(output_path)
    except Exception as e:
        raise IOError(f"Failed to save styled Excel file {output_path}: {e}")


def main():
    argv = sys.argv[1:]
    if len(argv) < 1:
        usage_and_exit("Missing arguments")

    config_path = argv[0]
    output_path = None
    if len(argv) >= 2:
        # If second argument provided and looks like .xlsx file, treat as explicit output filename
        maybe_output = argv[1]
        if maybe_output.lower().endswith((".xlsx", ".xls")):
            output_path = maybe_output

    # Load JSON config
    try:
        env_ordered_map = load_json_config(config_path)
    except Exception as e:
        usage_and_exit(str(e))

    env_names = list(env_ordered_map.keys())
    env_paths = list(env_ordered_map.values())

    # Validate YAML files & flatten
    env_flat_items_list: List[List[Tuple[str, Any]]] = []
    env_key_value_maps: Dict[str, Dict[str, Any]] = OrderedDict()
    try:
        for env, path in env_ordered_map.items():
            if not os.path.isfile(path):
                raise FileNotFoundError(f"YAML file for environment '{env}' not found at path: {path}")
            raw = read_yaml_file(path)
            flat_items = flatten_yaml(raw)  # list of (key, value) preserving order
            env_flat_items_list.append(flat_items)
            env_key_value_maps[env] = flatten_items_to_dict(flat_items)
    except Exception as e:
        usage_and_exit(str(e))

    # Build master key order (keys from first file in order, then appended new keys)
    master_keys = build_master_key_order(env_flat_items_list)

    # Build DataFrame
    df = build_dataframe(master_keys, env_names, env_key_value_maps)

    # Determine output filename if not provided
    if not output_path:
        output_path = autogen_output_filename(env_names)

    # Ensure output directory exists (if path contains directories)
    out_dir = os.path.dirname(os.path.abspath(output_path))
    if out_dir and not os.path.exists(out_dir):
        try:
            os.makedirs(out_dir, exist_ok=True)
        except Exception as e:
            usage_and_exit(f"Failed to create output directory {out_dir}: {e}")

    # Write Excel and apply highlights
    try:
        write_excel_with_highlight(df, output_path, env_names)
        print(f"Comparison Excel created successfully: {output_path}")
    except Exception as e:
        usage_and_exit(str(e))


if __name__ == "__main__":
    main()
